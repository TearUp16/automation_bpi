import io
import re
import calendar
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TABLE_STATUSES: List[str] = [
    "ALL NEGATIVE",
    "CALL OUTS",
    "EMAIL",
    "FLD VST",
    "SELF CURED",
]

MONTH_NAME_TO_NUM = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}

STATUS_ALIASES = {
    "CALL OUT": "CALL OUTS",
    "CALL-OUTS": "CALL OUTS",
    "FIELD VISIT": "FLD VST",
    "FLD VISIT": "FLD VST",
    "SELF-CURED": "SELF CURED",
}

OVERALL_START = 1
PTP_START     = 15
CURED_START   = 29
BLOCK_WIDTH   = 13

# Chart color constants
OVERALL_BAR_COLOR = "FF0000"   # red
OVERALL_LINE_COLOR = "FFFF00"  # yellow

PTP_BAR_COLOR = "4F81BD"       # blue
PTP_LINE_COLOR = "FF0000"      # red

CURED_BAR_COLOR = "4F81BD"     # blue
CURED_LINE_COLOR = "FF0000"    # red

VARIANCE_POSITIVE_BAR_COLOR = "4F81BD"  # blue
VARIANCE_NEGATIVE_BAR_COLOR = "D9EAF7"  # light blue


# ─────────────────────────────────────────────────────────────────────────────
# Data helpers
# ─────────────────────────────────────────────────────────────────────────────

def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip().upper())


def normalize_status(value: object) -> str:
    text = normalize_text(value)
    return STATUS_ALIASES.get(text, text)


def resolve_column(
    df: pd.DataFrame,
    *,
    by_name: Optional[str] = None,
    by_excel_index: Optional[int] = None,
) -> str:
    if by_name:
        target = normalize_text(by_name)
        for col in df.columns:
            if normalize_text(col) == target:
                return col
    if by_excel_index is not None:
        idx = by_excel_index - 1
        if 0 <= idx < len(df.columns):
            return df.columns[idx]
    raise KeyError(f"Column not found: {by_name or f'Excel #{by_excel_index}'}")


@st.cache_data(show_spinner=False)
def load_workbook_sheets(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
    return {
        s: pd.read_excel(io.BytesIO(file_bytes), sheet_name=s)
        for s in excel_file.sheet_names
    }


def extract_month_number(value: object) -> Optional[int]:
    if pd.isna(value):
        return None
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.notna(dt):
            return int(dt.month)
    except Exception:
        pass

    text = normalize_text(value)
    if not text:
        return None

    if text in MONTH_NAME_TO_NUM:
        return MONTH_NAME_TO_NUM[text]

    for name, num in MONTH_NAME_TO_NUM.items():
        if re.search(rf"\b{name}\b", text):
            return num

    m = re.fullmatch(r"(\d{1,2})(?:\.0+)?", text)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 12:
            return n
    return None


def month_num_to_name(n: int) -> str:
    return calendar.month_name[n].upper()


def get_month_span_label(df: pd.DataFrame) -> str:
    cutoff_col = resolve_column(df, by_name="CUT OFF MONTH", by_excel_index=8)
    months = sorted(set(
        df[cutoff_col].map(extract_month_number).dropna().astype(int).tolist()
    ))
    if not months:
        return "ALL MONTHS"
    if len(months) == 1:
        return month_num_to_name(months[0])
    return f"{month_num_to_name(min(months))} TO {month_num_to_name(max(months))}"


def get_detected_months(df: pd.DataFrame) -> List[int]:
    cutoff_col = resolve_column(df, by_name="CUT OFF MONTH", by_excel_index=8)
    return sorted(set(
        df[cutoff_col].map(extract_month_number).dropna().astype(int).tolist()
    ))


def filter_by_cutoff_month(df: pd.DataFrame, target_month: int) -> pd.DataFrame:
    cutoff_col = resolve_column(df, by_name="CUT OFF MONTH", by_excel_index=8)
    working = df.copy()
    working["_m"] = working[cutoff_col].map(extract_month_number)
    return working[working["_m"] == target_month].copy()


def filter_ptp(df: pd.DataFrame) -> pd.DataFrame:
    col = resolve_column(df, by_name="REMARKS (PTP/NO PTP)", by_excel_index=61)
    return df[df[col].apply(normalize_text) == "PTP"].copy()


def filter_cured(df: pd.DataFrame) -> pd.DataFrame:
    col = resolve_column(df, by_name="FINAL STATUS", by_excel_index=11)
    return df[df[col].apply(normalize_text) == "CURED"].copy()


def _cycle_sort_key(value: object) -> Tuple[float, str]:
    text = str(value).strip()
    m = re.search(r"(\d+)", text)
    if m:
        return (float(m.group(1)), text)
    return (float("inf"), text)


def _build_table(df: pd.DataFrame, is_sub: bool) -> pd.DataFrame:
    cycle_col   = resolve_column(df, by_name="CYCLE",                    by_excel_index=2)
    ob_col      = resolve_column(df, by_name="OB",                       by_excel_index=5)
    contact_col = resolve_column(df, by_name="CONTACT SOURCE (OVERALL)", by_excel_index=55)

    count_lbl = "NO. OF CASES" if is_sub else "COUNT OF ACCOUNT CYCLE"
    ob_lbl    = "OB"           if is_sub else "OB PER CYCLE"

    working = df.copy()
    working[cycle_col]   = working[cycle_col].fillna("").astype(str).str.strip()
    working[contact_col] = working[contact_col].map(normalize_status)
    working[ob_col]      = pd.to_numeric(working[ob_col], errors="coerce").fillna(0)
    working = working[working[cycle_col].ne("")].copy()

    cycle_order = (
        working[[cycle_col]]
        .drop_duplicates()
        .assign(_n=lambda x: x[cycle_col].astype(str).str.extract(r"(\d+)").astype(float))
        .sort_values(["_n", cycle_col])[cycle_col]
        .tolist()
    )

    rows: List[Dict] = []
    for cycle in cycle_order:
        cdf = working[working[cycle_col] == cycle]
        row: Dict = {
            ("", "Cycle"): cycle,
            ("TOTAL", "TOTAL COUNT OF CASES"): int(len(cdf)),
            ("TOTAL", "TOTAL OB"): float(cdf[ob_col].sum()),
        }
        for status in TABLE_STATUSES:
            sdf = cdf[cdf[contact_col] == status]
            row[(status, count_lbl)] = int(len(sdf))
            row[(status, ob_lbl)]    = float(sdf[ob_col].sum())
        rows.append(row)

    tbl = pd.DataFrame(rows)
    tbl.columns = pd.MultiIndex.from_tuples(tbl.columns)

    gt: Dict = {
        ("", "Cycle"): "Grand Total",
        ("TOTAL", "TOTAL COUNT OF CASES"): tbl[("TOTAL", "TOTAL COUNT OF CASES")].sum(),
        ("TOTAL", "TOTAL OB"): tbl[("TOTAL", "TOTAL OB")].sum(),
    }

    for col in tbl.columns:
        if col not in [("", "Cycle"), ("TOTAL", "TOTAL COUNT OF CASES"), ("TOTAL", "TOTAL OB")]:
            gt[col] = tbl[col].sum()

    return pd.concat([tbl, pd.DataFrame([gt])], ignore_index=True)


def build_summary_table(df: pd.DataFrame) -> pd.DataFrame:
    return _build_table(df, is_sub=False)


def build_sub_summary_table(df: pd.DataFrame) -> pd.DataFrame:
    return _build_table(df, is_sub=True)


def build_percentage_table(summary: pd.DataFrame, is_sub: bool = False) -> pd.DataFrame:
    count_key = "NO. OF CASES" if is_sub else "COUNT OF ACCOUNT CYCLE"
    ob_key    = "OB"           if is_sub else "OB PER CYCLE"

    detail = summary[summary[("", "Cycle")].astype(str).str.upper() != "GRAND TOTAL"].copy()
    rows: List[Dict] = []

    for _, row in detail.iterrows():
        tc = row[("TOTAL", "TOTAL COUNT OF CASES")]
        to = row[("TOTAL", "TOTAL OB")]

        pr: Dict = {
            ("", "Cycle"): row[("", "Cycle")],
            ("TOTAL", "TOTAL COUNT %"): 1 if tc else 0,
            ("TOTAL", "TOTAL OB %"): 1 if to else 0,
        }

        for s in TABLE_STATUSES:
            pr[(s, "COUNT %")] = (row[(s, count_key)] / tc) if tc else 0
            pr[(s, "OB %")]    = (row[(s, ob_key)]    / to) if to else 0

        rows.append(pr)

    pct = pd.DataFrame(rows)
    pct.columns = pd.MultiIndex.from_tuples(pct.columns)

    gt_row = summary[summary[("", "Cycle")].astype(str).str.upper() == "GRAND TOTAL"].iloc[0]
    gtc = gt_row[("TOTAL", "TOTAL COUNT OF CASES")]
    gto = gt_row[("TOTAL", "TOTAL OB")]

    gtr: Dict = {
        ("", "Cycle"): "Grand Total",
        ("TOTAL", "TOTAL COUNT %"): 1 if gtc else 0,
        ("TOTAL", "TOTAL OB %"): 1 if gto else 0,
    }

    for s in TABLE_STATUSES:
        gtr[(s, "COUNT %")] = (gt_row[(s, count_key)] / gtc) if gtc else 0
        gtr[(s, "OB %")]    = (gt_row[(s, ob_key)]    / gto) if gto else 0

    return pd.concat([pct, pd.DataFrame([gtr])], ignore_index=True)


def to_flat_preview(df: pd.DataFrame) -> pd.DataFrame:
    preview = df.copy()
    preview.columns = [
        sub if top in ("", "TOTAL") else f"{top} - {sub}"
        for top, sub in preview.columns
    ]
    return preview


# ─────────────────────────────────────────────────────────────────────────────
# Variance helpers
# ─────────────────────────────────────────────────────────────────────────────

def _extract_status_count_map(summary: Optional[pd.DataFrame], status: str) -> Dict[str, int]:
    if summary is None or summary.empty:
        return {}

    detail = summary[summary[("", "Cycle")].astype(str).str.upper() != "GRAND TOTAL"].copy()
    out: Dict[str, int] = {}

    for _, row in detail.iterrows():
        cycle = str(row[("", "Cycle")])
        out[cycle] = int(row[(status, "NO. OF CASES")])

    return out


def build_variance_rows(
    ptp_summary: Optional[pd.DataFrame],
    cured_summary: Optional[pd.DataFrame],
    status: str,
) -> List[List[object]]:
    """
    Variance = CURED Count - PTP Count
    Positive => CURED higher
    Negative => PTP higher
    """
    ptp_map = _extract_status_count_map(ptp_summary, status)
    cured_map = _extract_status_count_map(cured_summary, status)

    cycles = sorted(set(ptp_map) | set(cured_map), key=_cycle_sort_key)

    rows: List[List[object]] = []
    for cycle in cycles:
        ptp_count = int(ptp_map.get(cycle, 0))
        cured_count = int(cured_map.get(cycle, 0))
        variance = cured_count - ptp_count
        rows.append([cycle, ptp_count, cured_count, variance])

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel styles/helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_styles() -> dict:
    thin = Side(style="thin", color="000000")
    return dict(
        fill_title=PatternFill("solid", fgColor="D9EAD3"),
        fill_red=PatternFill("solid", fgColor="C00000"),
        fill_white=PatternFill("solid", fgColor="FFFFFF"),
        fill_green=PatternFill("solid", fgColor="C6EFCE"),
        fill_yellow=PatternFill("solid", fgColor="FFF2CC"),
        fill_blue=PatternFill("solid", fgColor="4F81BD"),
        fill_light_blue=PatternFill("solid", fgColor="D9EAF7"),
        border=Border(left=thin, right=thin, top=thin, bottom=thin),
        font_title=Font(name="Arial", size=12, bold=True, color="000000"),
        font_hdr=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        font_body=Font(name="Arial", size=10, bold=False, color="000000"),
        font_total=Font(name="Arial", size=10, bold=True, color="000000"),
        center=Alignment(horizontal="center", vertical="center"),
        left=Alignment(horizontal="left", vertical="center"),
        right=Alignment(horizontal="right", vertical="center"),
    )


def _apply(cell, fill, font, alignment, border, number_format=None):
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if number_format:
        cell.number_format = number_format


def _col(n: int) -> str:
    return get_column_letter(n)


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard writing
# ─────────────────────────────────────────────────────────────────────────────

def _write_summary_side_by_side(
    ws,
    overall_summary: Optional[pd.DataFrame],
    ptp_summary: Optional[pd.DataFrame],
    cured_summary: Optional[pd.DataFrame],
    overall_title: str,
    ptp_title: str,
    cured_title: str,
    start_row: int,
    s: dict,
) -> int:
    tables = [
        (overall_summary, overall_title, OVERALL_START, False),
        (ptp_summary,     ptp_title,     PTP_START,     True),
        (cured_summary,   cured_title,   CURED_START,   True),
    ]

    max_data_rows = 0
    for tbl, title, col_start, is_sub in tables:
        if tbl is None:
            continue

        count_lbl = "NO. OF CASES" if is_sub else "COUNT OF ACCOUNT CYCLE"
        ob_lbl    = "OB"           if is_sub else "OB PER CYCLE"
        c         = col_start

        title_start = _col(c)
        title_end   = _col(c + BLOCK_WIDTH - 1)
        ws.merge_cells(f"{title_start}{start_row}:{title_end}{start_row}")
        ws[f"{title_start}{start_row}"] = title
        _apply(ws[f"{title_start}{start_row}"],
               s["fill_title"], s["font_title"], s["left"], s["border"])

        hr1 = start_row + 1
        hr2 = start_row + 2

        ws.merge_cells(f"{_col(c)}{hr1}:{_col(c)}{hr2}")
        ws.merge_cells(f"{_col(c+1)}{hr1}:{_col(c+1)}{hr2}")
        ws.merge_cells(f"{_col(c+2)}{hr1}:{_col(c+2)}{hr2}")
        ws.merge_cells(f"{_col(c+3)}{hr1}:{_col(c+4)}{hr1}")
        ws.merge_cells(f"{_col(c+5)}{hr1}:{_col(c+6)}{hr1}")
        ws.merge_cells(f"{_col(c+7)}{hr1}:{_col(c+8)}{hr1}")
        ws.merge_cells(f"{_col(c+9)}{hr1}:{_col(c+10)}{hr1}")
        ws.merge_cells(f"{_col(c+11)}{hr1}:{_col(c+12)}{hr1}")

        ws.cell(hr1, c,    "Cycle")
        ws.cell(hr1, c+1,  "TOTAL COUNT OF CASES")
        ws.cell(hr1, c+2,  "TOTAL OB")
        ws.cell(hr1, c+3,  "ALL NEGATIVE")
        ws.cell(hr1, c+5,  "CALL OUTS")
        ws.cell(hr1, c+7,  "EMAIL")
        ws.cell(hr1, c+9,  "FLD VST")
        ws.cell(hr1, c+11, "SELF CURED")

        for cell in ws.iter_rows(min_row=hr1, max_row=hr1,
                                  min_col=c, max_col=c+12, values_only=False):
            for cl in cell:
                _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

        for offset, text in enumerate([count_lbl, ob_lbl] * 5, start=3):
            ws.cell(hr2, c + offset, text)
        for cell in ws.iter_rows(min_row=hr2, max_row=hr2,
                                  min_col=c+3, max_col=c+12, values_only=False):
            for cl in cell:
                _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

        ws.cell(hr2, c).border    = s["border"]
        ws.cell(hr2, c+1).border  = s["border"]
        ws.cell(hr2, c+2).border  = s["border"]

        count_offsets = [1, 3, 5, 7, 9, 11]
        ob_offsets    = [2, 4, 6, 8, 10, 12]
        data_start    = hr2 + 1

        for i, (_, row) in enumerate(tbl.iterrows(), start=data_start):
            label    = str(row[("", "Cycle")])
            is_total = normalize_text(label) == "GRAND TOTAL"
            bfont    = s["font_total"] if is_total else s["font_body"]

            values = [
                label,
                int(row[("TOTAL", "TOTAL COUNT OF CASES")]),
                float(row[("TOTAL", "TOTAL OB")]),
                int(row[("ALL NEGATIVE", count_lbl)]), float(row[("ALL NEGATIVE", ob_lbl)]),
                int(row[("CALL OUTS",    count_lbl)]), float(row[("CALL OUTS",    ob_lbl)]),
                int(row[("EMAIL",        count_lbl)]), float(row[("EMAIL",        ob_lbl)]),
                int(row[("FLD VST",      count_lbl)]), float(row[("FLD VST",      ob_lbl)]),
                int(row[("SELF CURED",   count_lbl)]), float(row[("SELF CURED",   ob_lbl)]),
            ]
            for offset, value in enumerate(values):
                cl = ws.cell(i, c + offset, value)
                _apply(cl, s["fill_white"], bfont,
                       s["left"] if offset == 0 else s["right"], s["border"])
                if offset in count_offsets:
                    cl.number_format = "#,##0"
                elif offset in ob_offsets:
                    cl.number_format = "#,##0"

        max_data_rows = max(max_data_rows, len(tbl))

    return start_row + 3 + max_data_rows


def _write_rate_side_by_side(
    ws,
    overall_pct: Optional[pd.DataFrame],
    ptp_pct: Optional[pd.DataFrame],
    cured_pct: Optional[pd.DataFrame],
    overall_title: str,
    ptp_title: str,
    cured_title: str,
    start_row: int,
    s: dict,
) -> int:
    tables = [
        (overall_pct, overall_title, OVERALL_START),
        (ptp_pct,     ptp_title,     PTP_START),
        (cured_pct,   cured_title,   CURED_START),
    ]

    max_data_rows = 0
    for tbl, title, col_start in tables:
        if tbl is None:
            continue
        c = col_start

        ws.merge_cells(f"{_col(c)}{start_row}:{_col(c+12)}{start_row}")
        ws[f"{_col(c)}{start_row}"] = title
        _apply(ws[f"{_col(c)}{start_row}"],
               s["fill_title"], s["font_title"], s["left"], s["border"])

        hr1 = start_row + 1
        hr2 = start_row + 2

        ws.merge_cells(f"{_col(c)}{hr1}:{_col(c)}{hr2}")
        ws.merge_cells(f"{_col(c+1)}{hr1}:{_col(c+1)}{hr2}")
        ws.merge_cells(f"{_col(c+2)}{hr1}:{_col(c+2)}{hr2}")
        ws.merge_cells(f"{_col(c+3)}{hr1}:{_col(c+4)}{hr1}")
        ws.merge_cells(f"{_col(c+5)}{hr1}:{_col(c+6)}{hr1}")
        ws.merge_cells(f"{_col(c+7)}{hr1}:{_col(c+8)}{hr1}")
        ws.merge_cells(f"{_col(c+9)}{hr1}:{_col(c+10)}{hr1}")
        ws.merge_cells(f"{_col(c+11)}{hr1}:{_col(c+12)}{hr1}")

        ws.cell(hr1, c,    "Cycle")
        ws.cell(hr1, c+1,  "TOTAL COUNT %")
        ws.cell(hr1, c+2,  "TOTAL OB %")
        ws.cell(hr1, c+3,  "ALL NEGATIVE")
        ws.cell(hr1, c+5,  "CALL OUTS")
        ws.cell(hr1, c+7,  "EMAIL")
        ws.cell(hr1, c+9,  "FLD VST")
        ws.cell(hr1, c+11, "SELF CURED")

        for cell in ws.iter_rows(min_row=hr1, max_row=hr1,
                                  min_col=c, max_col=c+12, values_only=False):
            for cl in cell:
                _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

        for offset, text in enumerate(["COUNT %", "OB %"] * 5, start=3):
            ws.cell(hr2, c + offset, text)
        for cell in ws.iter_rows(min_row=hr2, max_row=hr2,
                                  min_col=c+3, max_col=c+12, values_only=False):
            for cl in cell:
                _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

        ws.cell(hr2, c).border    = s["border"]
        ws.cell(hr2, c+1).border  = s["border"]
        ws.cell(hr2, c+2).border  = s["border"]

        data_start = hr2 + 1
        for i, (_, row) in enumerate(tbl.iterrows(), start=data_start):
            label    = str(row[("", "Cycle")])
            is_total = normalize_text(label) == "GRAND TOTAL"
            bfont    = s["font_total"] if is_total else s["font_body"]

            ws.cell(i, c, label)
            _apply(ws.cell(i, c), s["fill_white"], bfont, s["left"], s["border"])

            values = [
                row[("TOTAL", "TOTAL COUNT %")],  row[("TOTAL", "TOTAL OB %")],
                row[("ALL NEGATIVE", "COUNT %")], row[("ALL NEGATIVE", "OB %")],
                row[("CALL OUTS",    "COUNT %")], row[("CALL OUTS",    "OB %")],
                row[("EMAIL",        "COUNT %")], row[("EMAIL",        "OB %")],
                row[("FLD VST",      "COUNT %")], row[("FLD VST",      "OB %")],
                row[("SELF CURED",   "COUNT %")], row[("SELF CURED",   "OB %")],
            ]
            for offset, value in enumerate(values, start=1):
                cl = ws.cell(i, c + offset, float(value))
                _apply(cl, s["fill_white"], bfont, s["right"], s["border"], "0.00%")

        max_data_rows = max(max_data_rows, len(tbl))

    return start_row + 3 + max_data_rows


# ─────────────────────────────────────────────────────────────────────────────
# Chart helpers
# ─────────────────────────────────────────────────────────────────────────────

def _extract_status_chart_rows(
    summary: Optional[pd.DataFrame],
    status: str,
    is_sub: bool,
) -> List[List[object]]:
    if summary is None:
        return []

    count_key = "NO. OF CASES" if is_sub else "COUNT OF ACCOUNT CYCLE"
    ob_key    = "OB"           if is_sub else "OB PER CYCLE"

    detail = summary[summary[("", "Cycle")].astype(str).str.upper() != "GRAND TOTAL"].copy()

    rows: List[List[object]] = []
    for _, row in detail.iterrows():
        rows.append([
            str(row[("", "Cycle")]),
            int(row[(status, count_key)]),
            float(row[(status, ob_key)]),
        ])
    return rows


def _add_combo_chart(
    ws,
    title: str,
    table_col_start: int,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    anchor: str,
    bar_color: str,
    line_color: str,
):
    cats = Reference(
        ws,
        min_col=table_col_start,
        max_col=table_col_start,
        min_row=data_start_row,
        max_row=data_end_row,
    )

    count_data = Reference(
        ws,
        min_col=table_col_start + 1,
        max_col=table_col_start + 1,
        min_row=header_row,
        max_row=data_end_row,
    )

    ob_data = Reference(
        ws,
        min_col=table_col_start + 2,
        max_col=table_col_start + 2,
        min_row=header_row,
        max_row=data_end_row,
    )

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = title
    bar.y_axis.title = "Count"
    bar.height = 10.5
    bar.width = 13.5
    bar.gapWidth = 60
    bar.legend.position = "tr"
    bar.legend.overlay = True
    bar.x_axis.tickLblPos = "nextTo"
    bar.x_axis.delete = False

    try:
        bar.x_axis.noMultiLvlLbl = True
    except Exception:
        pass

    bar.add_data(count_data, titles_from_data=True)
    bar.set_categories(cats)

    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = True
    bar.dLbls.showCatName = False
    bar.dLbls.showSerName = False
    bar.dLbls.showLegendKey = False
    try:
        bar.dLbls.position = "ctr"
    except Exception:
        pass

    if bar.ser:
        bar.ser[0].graphicalProperties.solidFill = bar_color
        bar.ser[0].graphicalProperties.line.solidFill = bar_color

    line = LineChart()
    line.y_axis.title = "OB"
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    line.height = 10.5
    line.width = 13.5

    line.add_data(ob_data, titles_from_data=True)

    try:
        line.set_categories(cats)
    except Exception:
        pass

    line.dLbls = DataLabelList()
    line.dLbls.showVal = True
    line.dLbls.showCatName = False
    line.dLbls.showSerName = False
    line.dLbls.showLegendKey = False
    try:
        line.dLbls.position = "t"
    except Exception:
        pass

    if line.ser:
        line.ser[0].graphicalProperties.line.solidFill = line_color
        line.ser[0].graphicalProperties.line.width = 28575

    bar += line
    ws.add_chart(bar, anchor)


def _add_variance_chart(
    ws,
    title: str,
    table_col_start: int,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    anchor: str,
):
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.height = 10.5
    chart.width = 13.5
    chart.gapWidth = 35
    chart.legend = None
    chart.x_axis.title = "Variance Count (CURED - PTP)"
    chart.y_axis.title = "Cycle"

    try:
        chart.y_axis.reverseOrder = True
    except Exception:
        pass

    data = Reference(
        ws,
        min_col=table_col_start + 3,
        max_col=table_col_start + 3,
        min_row=header_row,
        max_row=data_end_row,
    )
    cats = Reference(
        ws,
        min_col=table_col_start,
        max_col=table_col_start,
        min_row=data_start_row,
        max_row=data_end_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = True
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    try:
        chart.dLbls.position = "outEnd"
    except Exception:
        pass

    if chart.ser:
        ser = chart.ser[0]
        ser.graphicalProperties.solidFill = VARIANCE_POSITIVE_BAR_COLOR
        ser.graphicalProperties.line.solidFill = VARIANCE_POSITIVE_BAR_COLOR
        ser.dPt = []

        for idx, row_num in enumerate(range(data_start_row, data_end_row + 1)):
            val = ws.cell(row=row_num, column=table_col_start + 3).value
            fill = VARIANCE_POSITIVE_BAR_COLOR if (val or 0) >= 0 else VARIANCE_NEGATIVE_BAR_COLOR

            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = fill
            pt.graphicalProperties.line.solidFill = fill
            ser.dPt.append(pt)

    ws.add_chart(chart, anchor)


def _write_chart_block_at(
    ws,
    block_title: str,
    chart_title: str,
    rows: List[List[object]],
    start_row: int,
    table_col_start: int,
    chart_anchor: str,
    s: dict,
    bar_color: str,
    line_color: str,
) -> None:
    title_start_col = table_col_start
    title_end_col = table_col_start + 2

    ws.merge_cells(f"{_col(title_start_col)}{start_row}:{_col(title_end_col)}{start_row}")
    ws.cell(start_row, title_start_col, block_title)
    _apply(ws.cell(start_row, title_start_col), s["fill_title"], s["font_title"], s["left"], s["border"])

    header_row = start_row + 1
    ws.cell(header_row, table_col_start,     "Cycle")
    ws.cell(header_row, table_col_start + 1, "Count")
    ws.cell(header_row, table_col_start + 2, "OB")

    for cell in ws.iter_rows(
        min_row=header_row, max_row=header_row,
        min_col=table_col_start, max_col=table_col_start + 2,
        values_only=False
    ):
        for cl in cell:
            _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

    data_start_row = header_row + 1

    if rows:
        for i, (cycle, count, ob) in enumerate(rows, start=data_start_row):
            c1 = ws.cell(i, table_col_start, cycle)
            c2 = ws.cell(i, table_col_start + 1, count)
            c3 = ws.cell(i, table_col_start + 2, ob)

            _apply(c1, s["fill_white"], s["font_body"], s["left"],  s["border"])
            _apply(c2, s["fill_white"], s["font_body"], s["right"], s["border"])
            _apply(c3, s["fill_white"], s["font_body"], s["right"], s["border"])

            c2.number_format = "#,##0"
            c3.number_format = "#,##0"

        data_end_row = data_start_row + len(rows) - 1
    else:
        data_end_row = data_start_row
        c1 = ws.cell(data_start_row, table_col_start, "No data")
        c2 = ws.cell(data_start_row, table_col_start + 1, 0)
        c3 = ws.cell(data_start_row, table_col_start + 2, 0)

        _apply(c1, s["fill_white"], s["font_body"], s["left"],  s["border"])
        _apply(c2, s["fill_white"], s["font_body"], s["right"], s["border"])
        _apply(c3, s["fill_white"], s["font_body"], s["right"], s["border"])

        c2.number_format = "#,##0"
        c3.number_format = "#,##0"

    _add_combo_chart(
        ws=ws,
        title=chart_title,
        table_col_start=table_col_start,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        anchor=chart_anchor,
        bar_color=bar_color,
        line_color=line_color,
    )


def _write_variance_block_at(
    ws,
    block_title: str,
    chart_title: str,
    rows: List[List[object]],
    start_row: int,
    table_col_start: int,
    chart_anchor: str,
    s: dict,
) -> None:
    title_start_col = table_col_start
    title_end_col = table_col_start + 3

    ws.merge_cells(f"{_col(title_start_col)}{start_row}:{_col(title_end_col)}{start_row}")
    ws.cell(start_row, title_start_col, block_title)
    _apply(
        ws.cell(start_row, title_start_col),
        s["fill_title"], s["font_title"], s["left"], s["border"]
    )

    header_row = start_row + 1
    headers = ["Cycle", "PTP Count", "CURED Count", "Variance"]
    for idx, hdr in enumerate(headers):
        ws.cell(header_row, table_col_start + idx, hdr)

    for cell in ws.iter_rows(
        min_row=header_row, max_row=header_row,
        min_col=table_col_start, max_col=table_col_start + 3,
        values_only=False
    ):
        for cl in cell:
            _apply(cl, s["fill_red"], s["font_hdr"], s["center"], s["border"])

    data_start_row = header_row + 1

    if rows:
        for i, (cycle, ptp_count, cured_count, variance) in enumerate(rows, start=data_start_row):
            values = [cycle, ptp_count, cured_count, variance]
            for offset, value in enumerate(values):
                cl = ws.cell(i, table_col_start + offset, value)

                if offset == 0:
                    _apply(cl, s["fill_white"], s["font_body"], s["left"], s["border"])
                else:
                    _apply(cl, s["fill_white"], s["font_body"], s["right"], s["border"])

                if offset >= 1:
                    cl.number_format = '#,##0;[Red]-#,##0'

        data_end_row = data_start_row + len(rows) - 1
    else:
        data_end_row = data_start_row
        default_vals = ["No data", 0, 0, 0]
        for offset, value in enumerate(default_vals):
            cl = ws.cell(data_start_row, table_col_start + offset, value)
            _apply(
                cl,
                s["fill_white"],
                s["font_body"],
                s["left"] if offset == 0 else s["right"],
                s["border"]
            )
            if offset >= 1:
                cl.number_format = '#,##0;[Red]-#,##0'

    _add_variance_chart(
        ws=ws,
        title=chart_title,
        table_col_start=table_col_start,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        anchor=chart_anchor,
    )


def _write_status_row_four_groups(
    ws,
    month_label: str,
    status: str,
    overall_summary: Optional[pd.DataFrame],
    ptp_summary: Optional[pd.DataFrame],
    cured_summary: Optional[pd.DataFrame],
    start_row: int,
    s: dict,
) -> int:
    overall_rows = _extract_status_chart_rows(overall_summary, status, is_sub=False)
    ptp_rows     = _extract_status_chart_rows(ptp_summary, status, is_sub=True)
    cured_rows   = _extract_status_chart_rows(cured_summary, status, is_sub=True)
    variance_rows = build_variance_rows(ptp_summary, cured_summary, status)

    block_height = max(
        max(len(overall_rows), len(ptp_rows), len(cured_rows), len(variance_rows)) + 8,
        24
    )

    _write_chart_block_at(
        ws=ws,
        block_title=f"OVERALL - {status}",
        chart_title=f"{month_label} - OVERALL - {status}",
        rows=overall_rows,
        start_row=start_row,
        table_col_start=1,
        chart_anchor=f"E{start_row}",
        s=s,
        bar_color=OVERALL_BAR_COLOR,
        line_color=OVERALL_LINE_COLOR,
    )

    _write_chart_block_at(
        ws=ws,
        block_title=f"PTP - {status}",
        chart_title=f"{month_label} - PTP - {status}",
        rows=ptp_rows,
        start_row=start_row,
        table_col_start=14,
        chart_anchor=f"R{start_row}",
        s=s,
        bar_color=PTP_BAR_COLOR,
        line_color=PTP_LINE_COLOR,
    )

    _write_chart_block_at(
        ws=ws,
        block_title=f"CURED - {status}",
        chart_title=f"{month_label} - CURED - {status}",
        rows=cured_rows,
        start_row=start_row,
        table_col_start=27,
        chart_anchor=f"AE{start_row}",
        s=s,
        bar_color=CURED_BAR_COLOR,
        line_color=CURED_LINE_COLOR,
    )

    _write_variance_block_at(
        ws=ws,
        block_title=f"VARIANCE - {status}",
        chart_title=f"{month_label} - VARIANCE - {status}",
        rows=variance_rows,
        start_row=start_row,
        table_col_start=40,
        chart_anchor=f"AS{start_row}",
        s=s,
    )

    return start_row + block_height + 2


def _write_month_chart_section(
    ws,
    month_label: str,
    overall_summary: Optional[pd.DataFrame],
    ptp_summary: Optional[pd.DataFrame],
    cured_summary: Optional[pd.DataFrame],
    start_row: int,
    s: dict,
) -> int:
    ws.merge_cells(f"A{start_row}:BC{start_row}")
    ws[f"A{start_row}"] = f"{month_label} CHARTS"
    _apply(ws[f"A{start_row}"], s["fill_title"], s["font_title"], s["left"], s["border"])

    group_row = start_row + 2

    ws.merge_cells(f"A{group_row}:M{group_row}")
    ws[f"A{group_row}"] = f"{month_label} - OVERALL RESPONSE BY CASES AND BALANCE"
    _apply(ws[f"A{group_row}"], s["fill_title"], s["font_title"], s["left"], s["border"])

    ws.merge_cells(f"N{group_row}:Z{group_row}")
    ws[f"N{group_row}"] = f"{month_label} - OVERALL RESPONSE BY CASES AND BALANCE - PTP"
    _apply(ws[f"N{group_row}"], s["fill_title"], s["font_title"], s["left"], s["border"])

    ws.merge_cells(f"AA{group_row}:AM{group_row}")
    ws[f"AA{group_row}"] = f"{month_label} - OVERALL RESPONSE BY CASES AND BALANCE - CURED"
    _apply(ws[f"AA{group_row}"], s["fill_title"], s["font_title"], s["left"], s["border"])

    ws.merge_cells(f"AN{group_row}:AR{group_row}")
    ws[f"AN{group_row}"] = f"{month_label} - VARIANCE (CURED COUNT - PTP COUNT)"
    _apply(ws[f"AN{group_row}"], s["fill_title"], s["font_title"], s["left"], s["border"])

    current_row = group_row + 2

    for status in TABLE_STATUSES:
        current_row = _write_status_row_four_groups(
            ws=ws,
            month_label=month_label,
            status=status,
            overall_summary=overall_summary,
            ptp_summary=ptp_summary,
            cured_summary=cured_summary,
            start_row=current_row,
            s=s,
        )

    return current_row + 2


def _build_charts_sheet(wb: Workbook, source_df: pd.DataFrame, s: dict) -> None:
    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    for col in range(1, 56):
        letter = _col(col)
        if col in (1, 14, 27, 40):
            ws.column_dimensions[letter].width = 16
        elif col in (2, 3, 15, 16, 28, 29, 41, 42, 43):
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 10

    ws.merge_cells("A1:BC1")
    ws["A1"] = "RESPONSE RATE CHARTS"
    _apply(ws["A1"], s["fill_title"], s["font_title"], s["left"], s["border"])

    ws.merge_cells("A2:BC2")
    ws["A2"] = (
        "Overall charts use red bars and yellow lines. "
        "PTP and CURED charts use blue bars and red lines. "
        "Variance charts use blue for positive bars and light blue for negative bars."
    )
    _apply(ws["A2"], s["fill_white"], s["font_body"], s["left"], s["border"])

    current_row = 4

    overall_summary = build_summary_table(source_df)
    ptp_all_df = filter_ptp(source_df)
    cured_all_df = filter_cured(source_df)

    ptp_all_summary = build_sub_summary_table(ptp_all_df) if not ptp_all_df.empty else None
    cured_all_summary = build_sub_summary_table(cured_all_df) if not cured_all_df.empty else None

    current_row = _write_month_chart_section(
        ws=ws,
        month_label="ALL MONTHS",
        overall_summary=overall_summary,
        ptp_summary=ptp_all_summary,
        cured_summary=cured_all_summary,
        start_row=current_row,
        s=s,
    )

    for month_num in get_detected_months(source_df):
        mname = month_num_to_name(month_num)
        month_df = filter_by_cutoff_month(source_df, month_num)
        ptp_df = filter_ptp(month_df)
        cured_df = filter_cured(month_df)

        month_summary = build_summary_table(month_df)
        ptp_summary = build_sub_summary_table(ptp_df) if not ptp_df.empty else None
        cured_summary = build_sub_summary_table(cured_df) if not cured_df.empty else None

        current_row = _write_month_chart_section(
            ws=ws,
            month_label=mname,
            overall_summary=month_summary,
            ptp_summary=ptp_summary,
            cured_summary=cured_summary,
            start_row=current_row,
            s=s,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Workbook builder
# ─────────────────────────────────────────────────────────────────────────────

def build_formatted_excel(source_df: pd.DataFrame, month_span_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    for c in range(1, 42):
        letter = _col(c)
        if c in (14, 28):
            ws.column_dimensions[letter].width = 2
        elif c in (1, 15, 29):
            ws.column_dimensions[letter].width = 22
        elif c in (2, 3, 16, 17, 30, 31):
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 16

    s = _make_styles()
    cur = 1

    overall_summary = build_summary_table(source_df)
    overall_pct     = build_percentage_table(overall_summary, is_sub=False)
    ptp_all         = filter_ptp(source_df)
    cured_all       = filter_cured(source_df)

    ptp_all_summary = build_sub_summary_table(ptp_all)   if not ptp_all.empty   else None
    cured_all_sum   = build_sub_summary_table(cured_all) if not cured_all.empty else None

    ptp_all_pct   = build_percentage_table(ptp_all_summary, is_sub=True) if ptp_all_summary is not None else None
    cured_all_pct = build_percentage_table(cured_all_sum,  is_sub=True) if cured_all_sum is not None else None

    cur = _write_summary_side_by_side(
        ws,
        overall_summary, ptp_all_summary, cured_all_sum,
        f"OVERALL RESPONSE BY CASES AND BALANCE ({month_span_label})",
        f"OVERALL RESPONSE BY CASES AND BALANCE - PTP ({month_span_label})",
        f"OVERALL RESPONSE BY CASES AND BALANCE - CURED ({month_span_label})",
        cur, s,
    )
    cur += 2

    cur = _write_rate_side_by_side(
        ws,
        overall_pct, ptp_all_pct, cured_all_pct,
        "OVERALL RESPONSE RATE",
        "OVERALL RESPONSE RATE - PTP",
        "OVERALL RESPONSE RATE - CURED",
        cur, s,
    )
    cur += 3

    for month_num in get_detected_months(source_df):
        mname         = month_num_to_name(month_num)
        month_df      = filter_by_cutoff_month(source_df, month_num)
        ptp_df        = filter_ptp(month_df)
        cured_df      = filter_cured(month_df)

        m_summary     = build_summary_table(month_df)
        m_pct         = build_percentage_table(m_summary, is_sub=False)
        ptp_summary   = build_sub_summary_table(ptp_df)   if not ptp_df.empty   else None
        cured_summary = build_sub_summary_table(cured_df) if not cured_df.empty else None
        ptp_pct       = build_percentage_table(ptp_summary, is_sub=True) if ptp_summary is not None else None
        cured_pct     = build_percentage_table(cured_summary, is_sub=True) if cured_summary is not None else None

        cur = _write_summary_side_by_side(
            ws,
            m_summary, ptp_summary, cured_summary,
            f"{mname} RESPONSE BY CASES AND BALANCE",
            f"{mname} RESPONSE BY CASES AND BALANCE - PTP",
            f"{mname} RESPONSE BY CASES AND BALANCE - CURED",
            cur, s,
        )
        cur += 2

        cur = _write_rate_side_by_side(
            ws,
            m_pct, ptp_pct, cured_pct,
            f"{mname} OVERALL RESPONSE RATE",
            f"{mname} OVERALL RESPONSE RATE - PTP",
            f"{mname} OVERALL RESPONSE RATE - CURED",
            cur, s,
        )
        cur += 3

    _build_charts_sheet(wb, source_df, s)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

def response_rate():
    st.title("BPI Reponse Rate Dashboard")

    uploaded_file = st.file_uploader(
        "PLEASE UPLOAD YOUR ENDORSEMENT WORKLIST(.xlsx)",
        type=["xlsx"]
    )

    if uploaded_file is None:
        st.info("Upload the workbook first.")
        st.stop()

    file_bytes = uploaded_file.getvalue()

    try:
        workbook = load_workbook_sheets(file_bytes)
    except Exception as exc:
        st.error(f"Unable to read the workbook: {exc}")
        st.stop()

    sheet_name = st.selectbox("Select source sheet", options=list(workbook.keys()), index=0)
    source_df = workbook[sheet_name]

    with st.expander("Preview source data", expanded=False):
        st.dataframe(source_df.head(20), use_container_width=True)

    try:
        month_span_label = get_month_span_label(source_df)
        detected_months = get_detected_months(source_df)
    except Exception as exc:
        st.error(f"Unable to parse month data: {exc}")
        st.stop()

    st.info(
        f"Detected months: **{', '.join(month_num_to_name(m) for m in detected_months)}** "
        f"({month_span_label})"
    )

    try:
        overall_summary = build_summary_table(source_df)
        overall_pct     = build_percentage_table(overall_summary, is_sub=False)
        ptp_all         = filter_ptp(source_df)
        cured_all       = filter_cured(source_df)
    except Exception as exc:
        st.error(f"Unable to build overall tables: {exc}")
        st.stop()

    st.subheader(f"Overall — Response by Cases and Balance ({month_span_label})")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.caption("Overall")
        st.dataframe(to_flat_preview(overall_summary), use_container_width=True, hide_index=True)
    with col2:
        st.caption("PTP")
        if not ptp_all.empty:
            st.dataframe(to_flat_preview(build_sub_summary_table(ptp_all)), use_container_width=True, hide_index=True)
        else:
            st.caption("No PTP data.")
    with col3:
        st.caption("CURED")
        if not cured_all.empty:
            st.dataframe(to_flat_preview(build_sub_summary_table(cured_all)), use_container_width=True, hide_index=True)
        else:
            st.caption("No CURED data.")

    st.subheader("Overall — Response Rate")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.caption("Overall")
        st.dataframe(to_flat_preview(overall_pct), use_container_width=True, hide_index=True)
    with col2:
        st.caption("PTP")
        if not ptp_all.empty:
            p = build_sub_summary_table(ptp_all)
            st.dataframe(to_flat_preview(build_percentage_table(p, is_sub=True)), use_container_width=True, hide_index=True)
        else:
            st.caption("No PTP data.")
    with col3:
        st.caption("CURED")
        if not cured_all.empty:
            p = build_sub_summary_table(cured_all)
            st.dataframe(to_flat_preview(build_percentage_table(p, is_sub=True)), use_container_width=True, hide_index=True)
        else:
            st.caption("No CURED data.")

    for month_num in detected_months:
        mname    = month_num_to_name(month_num)
        month_df = filter_by_cutoff_month(source_df, month_num)
        ptp_df   = filter_ptp(month_df)
        cured_df = filter_cured(month_df)

        st.markdown(f"---\n### {mname}")

        try:
            m_summary = build_summary_table(month_df)
            m_pct     = build_percentage_table(m_summary, is_sub=False)

            st.subheader(f"{mname} — Response by Cases and Balance")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.caption("Overall")
                st.dataframe(to_flat_preview(m_summary), use_container_width=True, hide_index=True)
            with c2:
                st.caption("PTP")
                if not ptp_df.empty:
                    st.dataframe(to_flat_preview(build_sub_summary_table(ptp_df)), use_container_width=True, hide_index=True)
                else:
                    st.caption("No PTP data.")
            with c3:
                st.caption("CURED")
                if not cured_df.empty:
                    st.dataframe(to_flat_preview(build_sub_summary_table(cured_df)), use_container_width=True, hide_index=True)
                else:
                    st.caption("No CURED data.")

            st.subheader(f"{mname} — Response Rate")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.caption("Overall")
                st.dataframe(to_flat_preview(m_pct), use_container_width=True, hide_index=True)
            with c2:
                st.caption("PTP")
                if not ptp_df.empty:
                    ps = build_sub_summary_table(ptp_df)
                    st.dataframe(to_flat_preview(build_percentage_table(ps, is_sub=True)), use_container_width=True, hide_index=True)
                else:
                    st.caption("No PTP data.")
            with c3:
                st.caption("CURED")
                if not cured_df.empty:
                    cs = build_sub_summary_table(cured_df)
                    st.dataframe(to_flat_preview(build_percentage_table(cs, is_sub=True)), use_container_width=True, hide_index=True)
                else:
                    st.caption("No CURED data.")

        except Exception as exc:
            st.error(f"Error building {mname} tables: {exc}")

    st.markdown("---")
    st.subheader("Download All Tables")

    try:
        excel_bytes = build_formatted_excel(source_df, month_span_label)
        st.download_button(
            label="⬇️ Download Response Rate",
            data=excel_bytes,
            file_name="BPI_Response_Rate.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        st.error(f"Unable to generate Excel file: {exc}")
